#coding=utf-8
import os
import sys
import xlrd
import datetime
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from flask import Flask,render_template,request,url_for,flash,redirect,jsonify,session
from new_db import DBConn
#from interface import transfer_excel_to_json
reload(sys)
sys.setdefaultencoding('utf8')

ALLOWED_EXTENSIONS = ['xlsx','xls']
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__),'uploads')
header_map = {
#all view
  '合同编号':'project_number',
  '项目名称':'project_name',
  '客户单位':'cust_organization',
  '客户名称':'cust_user',
  '项目状态':'status',
  '创建时间':'create_time',
#all have
  '序号':'id',
  '项目编号':'project_id',
  '样品名称':'sample_name',
  '样品编号':'sample_id',
  'OD260/280': 'od_260_or_280',
  'OD260/230': 'od_260_or_230',
  '备注': 'comment',
  '时间': 'time',
  '浓度(ng/ul)*': 'concentration',
  '体积(ul)*': 'volume',
  '总量(ug)*': 'quantum',
#send_sample_table
  '物种':'species',
  '提取部位':'extract_part',
  '管数':'sample_number',
#quality_inspection_table
  '25S/18S*':'25S_or_18S',
  'RIN':'rin',
  '建库类型':'database_type',
  '判定结果':'judgeresult',
#up_machine
  '上机类型':'upmachine_type',
  '上机模式':'upmachine_mode',
  '上机数据量':'up_quant',
#down_machine
  '下机数据量':'down_quant',
  'Reads':'reads',
  'Q30':'q30',
#return_sample
  '剩余量':'surplus'
}

title_list_dict = {'send_sample_table':['项目编号','样品名称','物种','提取部位','管数','备注','时间'],
                   'quality_inspection_table':['项目编号','样品名称','样品编号','OD260/230','OD260/280','25S/18S*','RIN',
                                               '体积','浓度','总量','建库类型','判定结果','备注','时间'],
                   'up_machine_table':['项目编号','样品名称','样品编号','上机类型','上机模式','上机数据量','备注','时间'],
                   'down_machine_table':['项目编号','样品名称','样品编号','下机数据量','Reads','Q30','备注','时间'],
                   'return_table':['项目编号','样品名称','物种','剩余量','备注','时间']}

def transfer_table_title(table_title):
    for index, title in enumerate(table_title):
        #if '数据量' in title:
        #    title = '数据量（raw data）'
        if str(title) in header_map.keys():
            table_title[index] = header_map[str(title)]

    return table_title

#transfer_excel_to_json
def transfer_excel_to_json2(file_name):
    data = xlrd.open_workbook(file_name)
    sheets = data.sheet_names()
    if len(sheets) != 1:
        return None

    table_data = []
    for sheet in sheets:
        table = data.sheet_by_name(sheet)
        n_rows = table.nrows
        n_cols = table.ncols
        table_title = table.row_values(0)
        table_title = transfer_table_title(table_title)
        for i in range(1,n_rows):
          table_data.append(dict(zip(table_title,table.row_values(i))))

    return table_data

def allow_files(filename):
    return '.' in filename and filename.rsplit('.',1)[1].lower() in ALLOWED_EXTENSIONS


def export_table_info(table_name, project_id, title_list):
    book = Workbook()
    sheet1 = book.worksheets[0]
    # connect database:
    db = DBConn()
    cmd = "select * from {table_name} where project_id = '{project_id}'".format(table_name=table_name,
                                                                                    project_id=project_id)
    results = db.execute(cmd)
    all_table_data = []
    for result in results:
        all_table_data.append(dict(result))

    for i, title in enumerate(title_list):
        sheet1.cell(row=1, column=i + 1).value = title
    for i, data in enumerate(all_table_data):
        for j, title in enumerate(title_list):
            sheet1.cell(row=i + 2, column=j + 1).value = data.get(header_map[title])
    export_path = os.path.join(os.path.abspath(os.path.dirname(__file__)),'static/export/')
    file_name = table_name + '_' + project_id + '_' + datetime.datetime.now().strftime("%Y%m%d%H%M%S") + '.xls'
    full_path_name = os.path.join(export_path, file_name)
    if not os.path.exists(export_path):
        os.mkdir(export_path)
        open(full_path_name,'w+').close()

    book.save(filename=full_path_name)

    return file_name

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['SECRET_KEY'] = 'hard to guess string'

@app.route('/')
def index():
    return render_template('send_sample_table.html')
'''
upload file part
'''
@app.route('/upload_send_sample_table',methods=['GET','POST'])
def upload_file():
    if request.method == 'POST':
        file = request.files['file']
        if file and allow_files(file.filename):
            filename = secure_filename(file.filename)
            if not os.path.exists(app.config['UPLOAD_FOLDER']):
                os.makedirs(app.config['UPLOAD_FOLDER'])
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            data = transfer_excel_to_json2(file_name=filepath)
            #input to database:
            db = DBConn()
            for each_record in data:
                db.insert('send_sample_table',each_record)
            return redirect(url_for('index'))
        elif file == '':
            flash('no select file!')
            return  redirect(url_for('index'))
        else:
            flash('only .xls or .xlsx file required!')
            return redirect(url_for('index'))

@app.route('/show_data',methods = ['POST','GET'])
def get_table_data():
    if request.method == 'POST':
        table_name = request.values.get('name',None)
        #分页
        page = request.values.get('page') if request.values.get('page') else 1
        rows = request.values.get('rows') if request.values.get('page') else 10
        offset = (int(page) - 1) * int(rows)
        #排序
        sort_col = request.values.get('sort') if request.values.get('sort') else 'sample_name'
        order_type = request.values.get('order') if request.values.get('order') else 'asc'
        if table_name:
            data = []
            i = 0
            db = DBConn()
            cmd = 'select * from {table} order by {sort_col} {order} limit {offset},{rows}'.format(
                table=table_name,
                sort_col=sort_col,
                order=order_type,
                offset=offset,
                rows=rows)
            results = db.execute(cmd)
            for result in results:
                data.append(dict(result))
                i += 1
            return jsonify({'total':i,'rows':data})
        else:
            return redirect(url_for('index'))
    return redirect(url_for('index'))

@app.route('/save_data',methods = ['GET','POST'])
def save_input_data():
    if request.method == 'POST':
        action = request.form.get('action',None)
        table_name = request.form.get('table_name',None)
        #just test two tables
        send_sample_table_dict = dict.fromkeys(['project_id','sample_name',
                                         'species','extract_part',
                                         'sample_number','comment','time'])
        quality_inspection_table_dict = dict.fromkeys(['project_id','sample_name','sample_id',
                                                       'od_260_or_230','od_260_or_280','25S_or_18S',
                                                       'rin','volume','concentration','quantum',
                                                       'database_type','judgeresult','comment','time'])
        up_machine_table_dict = dict.fromkeys(['project_id','sample_name',
                                               'upmachine_type','upmachine_mode','up_quant',
                                               'comment','time'])
        down_machine_table_dict = dict.fromkeys(['project_id','sample_name','sample_id','down_quant',
                                                 'reads','q30','comment','time'])
        return_table_dict = dict.fromkeys(['project_id','sample_name','species','surplus','comment','time'])

        all_table_dicts = dict(send_sample_table=send_sample_table_dict,
                               quality_inspection_table=quality_inspection_table_dict,
                               up_machine_table=up_machine_table_dict,
                               down_machine_table=down_machine_table_dict,
                               return_table=return_table_dict)

        send_sample_row = None
        for each_table,table_content in all_table_dicts.items():
            if table_name == each_table:
                send_sample_row = all_table_dicts[each_table]
                for key,value in send_sample_row.items():
                    send_sample_row[key] = request.form.get(key,None)

        db = DBConn()
        if action == 'insert' and send_sample_row:
            db.insert(table_name, send_sample_row)
            return jsonify({'success': 'true', 'errMsg': 'error'})
        elif action == 'update' and send_sample_row:
            try:
                db.update(table_name,
                          condition_dict={'sample_id':send_sample_row['sample_id']},
                          update_dict=send_sample_row)
            except Exception,e:
                db.update(table_name,
                          condition_dict={'sample_name': send_sample_row['sample_name']},
                          update_dict=send_sample_row)

            return jsonify({'success': 'true', 'errMsg': 'error'})
        else:
            return jsonify({'success': 'false', 'errMsg': 'error'})
    else:
        return redirect(url_for('index'))

'''
db = DBConn()
db.update(session.get('table_name'),
condition_dict={'sample_name': send_sample_row['sample_name']},
update_dict=send_sample_row)
'''

@app.route('/destroy_sample')
def del_select_data():
    sample_name = request.args.get('id')
    sample_time = request.args.get('time')
    table_name = request.args.get('table')
    db = DBConn()
    ans = db.delete(table_name,condition_dict={'sample_name':sample_name})
    if ans:
        return jsonify({'success':'true','errMsg':'error'})
    else:
        return jsonify({'success':'false','errMsg':'del fail!'})

@app.route('/down_table_file')
def export_table():
    table_name = request.args.get('table_name')
    project_id = request.args.get('project_id')
    title_list = title_list_dict[table_name]
    file_name = export_table_info(table_name,project_id,title_list)
    return file_name



if __name__ == '__main__':
    app.run()
