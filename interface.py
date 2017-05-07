#coding=utf-8
import datetime
import time
import os
import xlrd
from openpyxl import Workbook

from db import DBConn

ALLOWED_EXTENSIONS = ['xlsx','xls']
header_map = {
#all view
  '合同编号':'project_number',
  '项目名称':'project_name',
  '客户单位':'cust_organization',
  '客户名称':'cust_user',
  '项目状态':'status',
  '创建时间':'create_time',
#all have
  '序号':'id',
  '项目编号':'project_id',
  '样品名称':'sample_name',
  '样品编号':'sample_id',
  'OD260/280': 'od_260_or_280',
  'OD260/230': 'od_260_or_230',
  '备注': 'comment',
  '时间': 'time',
  '浓度(ng/ul)*': 'volume',
  '体积(ul)*': 'concentration',
  '总量(ug)*': 'quantum',
#send_sample_table
  '物种':'species',
  '提取部位':'extract_part',
  '管数':'sample_number',
#quality_inspection_table
  '25S/18S*':'25S_or_18S',
  'RIN':'rin',
  '建库类型':'database_type',
  '判定结果':'judgeresult',
#up_machine
  '上机类型':'upmachine_type',
  '上机模式':'upmachine_mode',
  '上机数据量':'up_quant',
#down_machine
  '下机数据量':'down_quant',
  'Reads':'reads',
  'Q30':'q30',
#return_sample
  '剩余量':'surplus'
}

def save_info(all_info, username, action='new'):
    db_instance = DBConn()
    try:
        project_id = save_sample_project_master_info(db_instance, all_info.get('sample_project_master_info'), username, action)
    except Exception, e:
        print e
        return {'data': '', 'errcode': 1, 'msg': '保存失败！项目编号已经存在，请另外选择项目编号。'}

    del all_info['sample_project_master_info']
    for key in all_info.keys():
        table_name = key.replace('_info', '')
        data_info = all_info.get(key)
        data_info['project_id'] = int(project_id)
        if action == 'update':
             db_instance.update(table_name, {'id': project_id}, data_info)
        else:
             db_instance.insert(table_name, data_info)

    msg = '更新成功！' if action == 'update' else '保存成功!'
    return {'data': '', 'errcode': 0, 'msg': msg}


def get_project_id_by_num(db_instance, project_number):
    cmd = "select id from sample_project_master where project_number=%s" % project_number
    result = db_instance.execute(cmd, get_all=False)

    return result[0]


def get_project_log_by_num(db_instance, project_number):
    cmd = "select project_log from sample_project_master where project_number=%s" % project_number
    result = db_instance.execute(cmd, get_all=False)

    return result[0] or ''

def save_table_info(db_instance, table_data, project_id, action):
    if action == 'update':
        cmd = "delete from sample_info_detail where project_id=%s" % project_id
        db_instance.execute(cmd)
    for row in table_data:
        del row['id']
        row['project_id'] = project_id
        if not row.get('sample_name'):
            continue
        else:
            db_instance.insert('sample_info_detail', row)


def save_sample_project_master_info(db_instance, data_info, username, action):
    time = datetime.datetime.now()
    if action == 'update':
        cmd = "select id from sample_project_master where project_number='%s'" % data_info['project_number']
        result = db_instance.execute(cmd, get_all=False)
        preject_id = result[0] #return id
        data_info['project_log'] += "\n%s: %s update this project.\n" % (time, username)
        db_instance.update('sample_project_master', {'id': preject_id}, data_info)
        return preject_id
    else:
        data_info['created_by'] = username
        data_info['create_time'] = datetime.datetime.now()
        data_info['project_log'] = "%s: %s create new project.\n" % (time, username)
        return db_instance.insert('sample_project_master', data_info)


def show_all_data(username, role='user'):
    data = []
    cmd = """select project_number,project_name,cust_organization,cust_user,status,create_time
            from sample_project_master spm where
          """
    if role == 'manager':
        cmd += "spm.project_leader='%s'" % username
    elif role == 'user':
        cmd += "spm.created_by='%s'" % username

    db = DBConn()
    result = db.execute(cmd)
    for i in result:
        temp_dict = dict(i)
        data.append(temp_dict)

    return data


def get_all_user_data():
    data = []
    cmd = "select username,e_mail email,tel,company,age,role,field,status,notes from user_info"
    results = DBConn().execute(cmd)
    for result in results:
        data.append(dict(result))

    return data


def get_one_project_data(project_number):
    #select multiple tables
    # why use project_id = spm.id: see save_info where data_info['project_id'] = int(project_id)
    cmd = """select * from sample_project_master spm,
             sample_species ss,
             sample_type st,
             dna_sample_sequencing_type dsst,
             rna_sample_sequencing_type rsst,
             sample_other other
             where ss.project_id=spm.id and
             st.project_id=spm.id and
             dsst.project_id=spm.id and
             rsst.project_id=spm.id and
             other.project_id=spm.id and
             spm.project_number='%s'""" %(project_number)
    db = DBConn()
    result = db.execute(cmd, get_all=False)
    data = dict(result)
    #dorp sample_info_detail table
    '''
    cmd = "select * from sample_info_detail where project_id=%s" % result['project_id']
    table_data = []
    results = db.execute(cmd)
    for result in results:
        table_data.append(dict(result))

    data['table_data'] = table_data
    '''
    return data



def get_analysis_data(username, role, selected_project):

    if not selected_project:
        return {}
    project_number = selected_project.split('-')[-1]
    cmd = """select a.* from analysis_master a,sample_project_master s
          where a.project_id=s.id and s.project_number=%s""" % project_number
    db = DBConn()
    result = db.execute(cmd, get_all=False)
    return dict(result) if result else {}


def phone_check(s):
    phone_prefix = ['130', '131', '132', '133', '134', '135', '136', '137', '138', '139', '150',
                    ' 151', '152', '153', '156', '158', '159', '170', '183', '182', '185', '186', '188', '189']

    return True if len(s) == 11 and s.isdigit() and s[:3] in phone_prefix else False


def get_user_role(username, password=''):
    db = DBConn()
    if phone_check(username):
        cmd = "select role,status from user_info where tel='%s'" % username
    else:
        cmd = "select role,status from user_info where username='%s'" % username
    if password:
        cmd += " and password='%s'" % password
    result = db.execute(cmd, get_all=False)
    role = result[0] if result else ''
    status = result[1] if result else ''
    return role, status


def get_other_info(username):
    db = DBConn()
    if phone_check(username):
        cmd = "select e_mail, tel, company, field, customer_name from user_info where tel='%s'" % username
    else:
        cmd = "select e_mail, tel, company, field, customer_name from user_info where username='%s'" % username
    result = db.execute(cmd, get_all=False)

    return result if result else ('', '', '', '')


def modify_base_info(info):
    time = datetime.datetime.now()
    info['update_time'] = time
    username = info.get('username')
    tel = info.get('tel')
    db = DBConn()
    cmd = "select id from user_info where username!='%s' and tel='%s'" % (username, tel)
    check_result = db.execute(cmd, get_all=False)
    if check_result:
        return {'data': '', 'errcode': 1, 'msg': '该电话号码已经被其它用户注册，请使用其它号码！'}
    else:
        db.update('user_info', {'username': username}, info)
        return {'data': '', 'errcode': 0, 'msg': '更新成功！'}


def change_password(info):
    username = info['username']
    db = DBConn()
    cmd = "select password from user_info where username='%s'" % username
    result = db.execute(cmd, get_all=False)
    if result and result[0] == info['old_passwd'] and info['new_passwd'] != info['old_passwd']:
        db.update('user_info', {'username': username}, {'password': info['new_passwd'],
                                                        'update_time': datetime.datetime.now()})
        return {'data': '', 'errcode': 0, 'msg': '更新成功！'}
    elif result and result[0] != info['old_passwd']:
        return {'data': '', 'errcode': 1, 'msg': '旧密码错误！'}
    elif info['new_passwd'] == info['old_passwd']:
        return {'data': '', 'errcode': 1, 'msg': '新密码和旧密码不能一样！'}
    else:
        return {'data': '', 'errcode': 1, 'msg': '用户名不存在！'}


def save_user_info(info):
    time = datetime.datetime.now()
    info['create_time'] = time
    info['update_time'] = time
    info['status'] = 'R'  # 'R' means review
    if len(info['password']) < 10:
        return {'data': '', 'errcode': 3, 'msg': "密码设置过短，请设置为10位以上！"}
    db = DBConn()
    name_result = db.execute("select id, status from user_info where username='%s'" % info.get('username'), get_all=False)
    tel_result = db.execute("select id, status from user_info where tel='%s'" % info.get('tel'), get_all=False)

    def fail_info(sql_result, check_type):
        status = sql_result[1]
        if status == 'R':
            msg = '该{check_type}之前已经注册，但状态处于等待中，请使用其它{check_type}注册！'
        elif status == 'Y':
            msg = '注册失败：该{check_type}已经注册，请使用其它{check_type}！'
        else:
            msg = '注册失败：该{check_type}已经存在！'

        return {'data': '', 'errcode': 2, 'msg': msg.format(check_type=check_type)}

    if name_result:
        ret = fail_info(name_result, '名字')
    elif tel_result:
        ret = fail_info(tel_result, '电话号码')
    else:
        db.insert('user_info', info)
        ret = {'data': '', 'errcode': 0, 'msg': '注册成功：等待审核！'}

    return ret


def get_detail_sample_data(project_number):
    data = []
    cmd = """SELECT d.* FROM sample_project_master m
            where m.id=d.project_id and m.project_number='%s'""" % project_number
    db = DBConn()
    results = db.execute(cmd)
    for result in results:
        data.append(dict(result))

    return data


def get_analysis_table_data(username, selected_project):
    data = {}
    db = DBConn()
    project_number = selected_project.split('-')[-1]
    project_id = get_project_id_by_num(db, project_number)
    cmd = """select info.* from sample_packet_information info, analysis_master m
                  where m.id=info.master_id and m.project_id=%s""" % project_id
    results = db.execute(cmd)
    data['sample_packet_information'] = [dict(i) for i in results]
    cmd = """select info.* from compare_table info, analysis_master m
                  where m.id=info.master_id and m.project_id=%s""" % project_id
    results = db.execute(cmd)
    data['compare_table'] = [dict(i) for i in results]

    return data


def save_status(project_number, status):
    db = DBConn()
    db.update('sample_project_master', {'project_number': project_number}, {'status': status})


def admin_save_user_info(username, status):
    db = DBConn()
    db.update('user_info', {'username': username}, {'status': status})


def get_project_number_list(username, user_role):
    if user_role == 'manager':
        cmd = "select concat(project_name, '-', project_number) from sample_project_master where project_leader='%s'" % username
    elif user_role == 'user':
        cmd = "select concat(project_name, '-', project_number) from sample_project_master where created_by='%s'" % username

    db = DBConn()
    results = db.execute(cmd)

    return [i[0] for i in results]


def get_manager_list():
    db = DBConn()
    cmd = "select username from user_info where role='manager'"

    result = db.execute(cmd)

    return [i[0] for i in result]

def transfer_table_title(table_title):
    for index, title in enumerate(table_title):
        if str(title) in header_map.keys():
            table_title[index] = header_map[str(title)]

    return table_title

def transfer_excel_to_json(file_name):
    # file_name = os.path.basename(file_name)
    data = xlrd.open_workbook(file_name)
    sheets = data.sheet_names()
    json_data = {}
    for sheet in sheets:
        table = data.sheet_by_name(sheet)
        n_rows = table.nrows
        n_cols = table.ncols
        table_data = []
        table_title = table.row_values(0)
        table_title = transfer_table_title(table_title)
        for i in range(1, n_rows):
            table_data.append(dict(zip(table_title, table.row_values(i))))
        json_data[sheet] = table_data

    return json_data


def upload_excel(io_stream):
    """
    Upload file and write it to database
    :param filename: filename
    :param io_stream: IO Stream
    :return: dict information
    """
    upload_path = os.path.dirname(__file__) + '/static/import/'
    file_path = os.path.join(upload_path, 'temp.xls')
    #juge file whether has exist
    if os.path.exists(file_path):
        os.remove(file_path)
    temp_file_path = file_path + '~'
    output_file = open(temp_file_path, 'wb')
    io_stream.seek(0)
    while True:
        data = io_stream.read(2 << 16)
        if not data:
            break
        output_file.write(data)

    output_file.flush()
    output_file.close()
    os.rename(temp_file_path, file_path)
    data = transfer_excel_to_json(file_path)
    return {'data': data, 'errcode': 0, 'msg': '上传成功'}


def call_sms_service(ipaddr):
    import requests
    account_sid = '518d949550f3567866c2694f4988f889'
    auth_token = 'ca4624f01d3ed40b30d09ff1cc1c374b'
    service_host = 'https://api.ucpaas.com/2014-06-30/Accounts/{account_sid}/Messages/templateSMS'
    payload = {
         "templateSMS": {
             "appId": "caabd9b5be1b4ff8a8592068b230c339",
             "param": "0000",
             "templateId": "30213",
             "to": ipaddr
             }
    }

    head = {
        'Accept': 'application/json',
        'Content-Type': 'application/json;charset=utf-8',
        'Authorization': 'ZTAzYmM5MTA2YzZlZDBlYWViZmNlOGMzNjhmZGNkNDg6MjAxNDA2MjMxODUwMjE'
    }

    r = requests.post(service_host, data=payload, headers=head)
    print r
    return

def export_table_info(table_name,project_id,title_list):
    book = Workbook()
    sheet1 = book.worksheets[0]
    #connect database:
    db = DBConn()
    cmd = "select * from {table_name} where project_id = '{project_id}'".format(table_name=table_name,project_id=project_id)
    results = db.execute(cmd)
    for i,title in enumerate(title_list):
        sheet1.cell(row=1,column=i+1).value = title
    for i,result in enumerate(results):
        for j,title in enumerate(title_list):
            sheet1.cell(row=i+2,column=j+1).value = result.get(header_map[title])
    export_path = os.path.dirname(__file__) + '/static/export/'
    file_name = table_name + '_' + project_id + '_' + datetime.datetime.now().strftime("%Y%m%d%H%M%S") + '.xls'
    full_path_name = os.path.join(export_path,file_name)
    book.save(filename=full_path_name)

    return file_name


def export_user_info():
    book = Workbook()
    sheet1 = book.worksheets[0]
    all_user_data = get_all_user_data()
    title_map = {'用户名': 'username',
                 '电子邮件': 'email',
                 '手机号码': 'tel',
                 '客户单位': 'company',
                 '年龄': 'age',
                 '角色': 'role',
                 '研究领域': 'field',
                 '状态': 'status',
                 '备注': 'notes'}
    title_list = ['用户名', '电子邮件', '手机号码', '客户单位', '年龄', '角色', '研究领域', '状态']
    for i, title in enumerate(title_list):
        sheet1.cell(row=1, column=i + 1).value = title
    for i, data in enumerate(all_user_data):
        for j, title in enumerate(title_list):
            sheet1.cell(row=i+2, column=j + 1).value = data.get(title_map[title])
    export_path = os.path.dirname(__file__) + '/static/export/'
    file_name = 'user_info_' + datetime.datetime.now().strftime("%Y%m%d%H%M%S") + '.xls'
    full_file_name = os.path.join(export_path, file_name)
    # remove old user info table
    for name in os.listdir(export_path):
        if 'user_info' in name:
            os.remove(os.path.join(export_path, name))
    book.save(filename=full_file_name)

    return file_name


def upload_project_file(io_stream, filename, project_number, project_name):
    filename = filename.split('\\')[-1]
    project_folder = project_number + '-' + project_name
    upload_path = os.path.join(os.path.dirname(__file__), 'static/import/', project_folder)
    if not os.path.exists(upload_path):
        os.makedirs(upload_path)
    file_path = os.path.join(upload_path, filename)
    # juge file whether has exist
    if os.path.exists(file_path):
        os.remove(file_path)
        print "delete had exist file: %s" % filename
    temp_file_path = file_path + '~'
    output_file = open(temp_file_path, 'wb')
    io_stream.seek(0)
    while True:
        data = io_stream.read(2 << 16)
        if not data:
            break
        output_file.write(data)
    output_file.flush()
    output_file.close()
    os.rename(temp_file_path, file_path)

    return {'data': '', 'errcode': 0, 'msg': "upload ok!"}


def get_project_files(project_number, project_name):
    file_list = []
    project_folder = project_number + '-' + project_name
    project_folder_path = os.path.join(os.path.dirname(__file__), 'static/import/', project_folder)
    if not os.path.exists(project_folder_path):
        return {'data': [], 'errcode': 0, 'msg': ""}
    files = os.listdir(project_folder_path)
    for file_name in files:
        file_path = os.path.join(project_folder_path, file_name)
        stat_info = os.stat(file_path)
        file_list.append({
            'file_name': file_name,
            'create_time': time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(stat_info.st_ctime)),
            'update_time': time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(stat_info.st_mtime))
        })

    return {'data': file_list, 'errcode': 0, 'msg': ""}
'''
add some new functions by chencheng on 2017-04-24
'''
def transfer_excel_to_json2(file_name):
    data = xlrd.open_workbook(file_name)
    sheets = data.sheet_names()
    if len(sheets) != 1:
        return None

    table_data = []
    for sheet in sheets:
        table = data.sheet_by_name(sheet)
        n_rows = table.nrows
        n_cols = table.ncols
        table_title = table.row_values(0)
        table_title = transfer_table_title(table_title)
        for i in range(1,n_rows):
          table_data.append(dict(zip(table_title,table.row_values(i))))

    return table_data

def allow_files(filename):
    return '.' in filename and filename.rsplit('.',1)[1].lower() in ALLOWED_EXTENSIONS

def check_table(table_title):
    if 'extract_part' in table_title:
        table_name = 'send_sample_table'
    elif 'judgeresult' in table_title:
        table_name = 'quality_inspection_table'
    elif 'up_quant' in table_title:
        table_name = 'up_machine_table'
    elif 'down_quant' in table_title:
        table_name = 'down_machine_table'
    elif 'surplus' in table_title:
        table_name = 'return_table'
    else:
        table_name = None
    return table_name


if __name__ == '__main__':
    # print transfer_excel_to_json('/home/chenjialin/下载/Table.1.xls')
    print get_project_files('111', '111')